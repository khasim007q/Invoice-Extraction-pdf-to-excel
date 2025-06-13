import os
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import string # Import string module for sanitization

def sanitize_excel_cell_value(value):
    """
    Sanitizes a value for writing to an Excel cell.
    Converts to string, removes non-printable characters, and replaces newlines.
    """
    if value is None:
        return ""
    
    s_value = str(value)
    # Remove characters that are not in string.printable (e.g., control characters)
    sanitized_s_value = ''.join(char for char in s_value if char in string.printable)
    # Replace common problematic characters or sequences
    sanitized_s_value = sanitized_s_value.replace('\r\n', ' ').replace('\n', ' ').replace('\r', ' ')
    sanitized_s_value = sanitized_s_value.strip()
    return sanitized_s_value

def write_to_excel(parsed_invoices, output_filepath, sheet_name="Invoices"):
    """
    Writes parsed invoice data to an Excel file.

    Args:
        parsed_invoices (list): A list of dictionaries, where each dictionary
                                represents a parsed invoice section.
                                Expected structure:
                                {
                                    "Invoice Type": "...",
                                    "Invoice Number": "...",
                                    "Order ID": "...",
                                    "Invoice Date": "...",
                                    "Total Amount": "...",
                                    "Items": [
                                        {"Description": "...", "Quantity": ..., "Unit Price": ..., "Total Item Price": ...},
                                        ...
                                    ]
                                }
        output_filepath (str): The full path to the output Excel file (e.g., "output/invoices.xlsx").
        sheet_name (str): The name of the sheet to write the data to.
    """
    if not parsed_invoices:
        print("Debug: No parsed invoice data to write to Excel.")
        return

    try:
        # Create a new workbook and select the active sheet
        wb = Workbook()
        
        # Ensure sheet_name is not too long for Excel (max 31 chars)
        # Use a safe slice or truncate if needed, but the main_extractor will handle shorter names
        ws = wb.active
        ws.title = sheet_name[:31] # Truncate to max 31 characters for safety
        print(f"Debug: Created workbook and sheet '{ws.title}'.")

        # Define header row (general invoice data first, then item specific)
        headers = [
            "Invoice Type", "Invoice Number", "Order ID", "Invoice Date", "Total Amount",
            "Item Description", "Item Quantity", "Item Unit Price", "Item Total Price"
        ]
        
        # Apply header styling
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        thin_border = Border(left=Side(style='thin'), 
                             right=Side(style='thin'), 
                             top=Side(style='thin'), 
                             bottom=Side(style='thin'))

        ws.append(headers)
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center", vertical="center")
        print("Debug: Wrote headers to Excel.")

        # Write data rows
        for invoice in parsed_invoices:
            # Get and sanitize base invoice details
            invoice_type = sanitize_excel_cell_value(invoice.get("Invoice Type"))
            invoice_number = sanitize_excel_cell_value(invoice.get("Invoice Number"))
            order_id = sanitize_excel_cell_value(invoice.get("Order ID"))
            invoice_date = sanitize_excel_cell_value(invoice.get("Invoice Date"))
            
            # Attempt to convert total_amount to float, otherwise keep as sanitized string
            total_amount_str = sanitize_excel_cell_value(invoice.get("Total Amount"))
            try:
                total_amount = float(total_amount_str)
            except ValueError:
                total_amount = total_amount_str # Keep as string if conversion fails

            base_row = [
                invoice_type,
                invoice_number,
                order_id,
                invoice_date,
                total_amount
            ]

            items = invoice.get("Items", [])
            if items:
                for i, item in enumerate(items):
                    row_data = list(base_row) # Create a copy of base_row
                    
                    # Sanitize item description
                    item_description = sanitize_excel_cell_value(item.get("Description"))

                    # Attempt to convert item numerical values to float/int
                    item_quantity_str = sanitize_excel_cell_value(item.get("Quantity"))
                    try:
                        item_quantity = int(float(item_quantity_str)) # Handle floats like "1.0"
                    except ValueError:
                        item_quantity = item_quantity_str

                    item_unit_price_str = sanitize_excel_cell_value(item.get("Unit Price"))
                    try:
                        item_unit_price = float(item_unit_price_str)
                    except ValueError:
                        item_unit_price = item_unit_price_str

                    item_total_price_str = sanitize_excel_cell_value(item.get("Total Item Price"))
                    try:
                        item_total_price = float(item_total_price_str)
                    except ValueError:
                        item_total_price = item_total_price_str

                    row_data.extend([
                        item_description,
                        item_quantity,
                        item_unit_price,
                        item_total_price
                    ])
                    ws.append(row_data)
                    # Apply border to all cells in the row
                    for col_idx in range(1, len(row_data) + 1):
                        ws.cell(row=ws.max_row, column=col_idx).border = thin_border
            else:
                # If no items, still append the main invoice data
                ws.append(list(base_row) + ["", "", "", ""]) # Add empty cells for item columns
                # Apply border to all cells in the row
                for col_idx in range(1, len(headers) + 1):
                    ws.cell(row=ws.max_row, column=col_idx).border = thin_border
            print(f"Debug: Appended invoice data for Order ID: {order_id}")

        # Adjust column widths for better readability
        for col_idx, header in enumerate(headers, 1):
            max_length = len(header)
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=col_idx, max_col=col_idx):
                for cell in row:
                    try:
                        # Consider string length for width calculation
                        cell_value_str = str(cell.value) if cell.value is not None else ""
                        max_length = max(max_length, len(cell_value_str))
                    except TypeError:
                        pass 
            adjusted_width = (max_length + 2)
            ws.column_dimensions[get_column_letter(col_idx)].width = adjusted_width
        print("Debug: Adjusted column widths.")

        # Save the workbook
        wb.save(output_filepath)
        print(f"Debug: Successfully saved workbook to '{output_filepath}'.")

    except ModuleNotFoundError:
        print("Error: 'openpyxl' library not found. Please install it using 'pip install openpyxl'")
    except Exception as e:
        print(f"Error during Excel writing process: {e}")
        import traceback
        traceback.print_exc() # Print full traceback for unexpected errors

# This block is for testing excel_writer.py independently if needed
if __name__ == "__main__":
    print("--- Running excel_writer.py independently for testing ---")
    
    # Create a dummy output folder for independent testing
    test_output_folder = "test_output_excel"
    os.makedirs(test_output_folder, exist_ok=True)
    test_output_filepath = os.path.join(test_output_folder, "test_invoices.xlsx")

    # Sample parsed invoice data
    sample_invoices = [
        {
            "Invoice Type": "Tax Invoice",
            "Invoice Number": "INV123",
            "Order ID": "ORD987",
            "Invoice Date": "01-01-2024",
            "Total Amount": "100.00",
            "Items": [
                {"Description": "Product A", "Quantity": 2, "Unit Price": 50.00, "Total Item Price": 100.00}
            ]
        },
        {
            "Invoice Type": "Credit Note",
            "Invoice Number": "CN456",
            "Order ID": "ORD654",
            "Invoice Date": "02-01-2024",
            "Total Amount": "-25.00",
            "Items": [
                {"Description": "Refund Fee", "Quantity": 1, "Unit Price": -25.00, "Total Item Price": -25.00}
            ]
        },
        {
            "Invoice Type": "Tax Invoice",
            "Invoice Number": "INV789",
            "Order ID": "ORD321",
            "Invoice Date": "03-01-2024",
            "Total Amount": "0.00",
            "Items": [] # No items
        }
    ]

    write_to_excel(sample_invoices, test_output_filepath, "Sample Invoices")
    print(f"Independent test complete. Check '{test_output_filepath}'")
